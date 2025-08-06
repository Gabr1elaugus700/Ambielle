import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from workflow.models import Cliente, Tarefa, Suporte

def export_relatorio_xlsx(request, relatorio_tipo):
    """
    Exporta relatórios em Excel (.xlsx) com cabeçalhos em negrito e filtragem.
    """
    data_inicial = request.GET.get("data_inicial")
    data_final = request.GET.get("data_final")
    cliente_nome = request.GET.get("cliente")
    status = request.GET.get("status")
    tipo_servico = request.GET.get("tipo_servico")
    formato = request.GET.get("formato", "excel")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{relatorio_tipo}.{formato}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Relatório - {relatorio_tipo.replace("_", " ").title()}'

    bold_font = Font(bold=True)

    def aplicar_filtros(queryset, modelo="tarefa"):
        if cliente_nome:
            if modelo == "cliente":
                queryset = queryset.filter(nome__icontains=cliente_nome)
            else:
                queryset = queryset.filter(cliente__nome__icontains=cliente_nome)
        if modelo in ("tarefa", "suporte"):
            data_field = "data_inicio" if modelo == "tarefa" else "data_suporte"
            if data_inicial:
                queryset = queryset.filter(**{f"{data_field}__gte": data_inicial})
            if data_final:
                queryset = queryset.filter(**{f"{data_field}__lte": data_final})
        if modelo == "tarefa":
            if status:
                queryset = queryset.filter(status=status)
            if tipo_servico:
                queryset = queryset.filter(tipo_servico__nome__icontains=tipo_servico)
        return queryset

    if relatorio_tipo == "clientes":
        headers = ['ID', 'Nome do Cliente', 'CNPJ', 'Endereço', 'Cidade', 'Telefone', 'Contato']
        ws.append(headers)
        for cell in ws[1]: cell.font = bold_font

        clientes = aplicar_filtros(Cliente.objects.all(), "cliente")
        for cliente in clientes:
            ws.append([cliente.id, cliente.nome, cliente.cnpj, cliente.endereco, cliente.bairro or "N/A", cliente.telefone, cliente.contato_principal])

    elif relatorio_tipo == "clientes_servicos":
        headers = ['ID', 'Cliente', 'CNPJ', 'Serviço', 'Status']
        ws.append(headers)
        for cell in ws[1]: cell.font = bold_font

        tarefas = aplicar_filtros(Tarefa.objects.all())
        for tarefa in tarefas:
            ws.append([tarefa.id, tarefa.cliente.nome, tarefa.cliente.cnpj, tarefa.tipo_servico.nome, tarefa.status])

    elif relatorio_tipo == "clientes_servicos_valores":
        headers = ['ID', 'Cliente', 'CNPJ', 'Serviço', 'Valor (R$)', 'Status']
        ws.append(headers)
        for cell in ws[1]: cell.font = bold_font

        tarefas = aplicar_filtros(Tarefa.objects.all())
        for tarefa in tarefas:
            ws.append([tarefa.id, tarefa.cliente.nome, tarefa.cliente.cnpj, tarefa.tipo_servico.nome, tarefa.valor_total_servico or "N/A", tarefa.status])

    elif relatorio_tipo == "clientes_suporte":
        headers = ['ID', 'Cliente', 'CNPJ', 'Descrição do Suporte', 'Valor Total (R$)']
        ws.append(headers)
        for cell in ws[1]: cell.font = bold_font

        suportes = aplicar_filtros(Suporte.objects.all(), "suporte")
        for suporte in suportes:
            ws.append([suporte.id, suporte.cliente.nome, suporte.cliente.cnpj, suporte.descricao, suporte.valor_total or "N/A"])

    elif relatorio_tipo == "demandas_por_cliente":
        headers = ['ID', 'Demanda/Serviço', 'Prazo Final', 'Data de Início', 'Cliente', 'CNPJ', 'Valor (R$)']
        ws.append(headers)
        for cell in ws[1]: cell.font = bold_font

        tarefas = aplicar_filtros(Tarefa.objects.all())
        for tarefa in tarefas:
            ws.append([tarefa.id, tarefa.tipo_servico.nome, tarefa.prazo_final.strftime('%d/%m/%Y') if tarefa.prazo_final else "N/A",
                       tarefa.data_inicio.strftime('%d/%m/%Y'), tarefa.cliente.nome, tarefa.cliente.cnpj, tarefa.valor_total_servico or "N/A"])

    wb.save(response)
    return response
